import os
import pandas as pd
import numpy as np 
from tkinter import Tk, filedialog, simpledialog, messagebox, ttk
import time
import subprocess
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference
import plotly.express as px
import logging
import threading
import sys
from openpyxl.chart.series import Series
import tkinter as tk


def setup_logging():
    logging.basicConfig(filename='merge_excel_log.txt', level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s')
    logging.info('Logging setup complete.')

def convert_and_clean_date_column(all_data):
    if 'Date' in all_data.columns:
        try:
            # Convert Date column to datetime format
            all_data['Date'] = pd.to_datetime(all_data['Date'], errors='coerce')

            # Fill any missing date values (forward and backward filling)
            all_data['Date'] = all_data['Date'].fillna(method='ffill')
            all_data['Date'] = all_data['Date'].fillna(method='bfill')

        except Exception as e:
            logging.error(f"Error processing Date column: {e}")
    return all_data


# 간단한 파일 이름 생성 함수
def generate_simple_filename(base_name, index):
    return f"{base_name}_{index}.xlsx"


def add_chart_to_excel(file_path):
    logging.info(f"Loading workbook: {file_path}")
    print(f"Loading workbook: {file_path}")
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        if ws is None or ws.max_row == 1 or ws.max_column == 1:
            logging.info("Worksheet is empty or not properly loaded.")
            print("Worksheet is empty or not properly loaded.")
            return

        logging.info("Creating scatter chart...")
        print("Creating scatter chart...")
        chart = ScatterChart()
        chart.title = "Scatter Plot"
        chart.style = 13
        chart.x_axis.title = "Line"
        chart.y_axis.title = "Values"

        logging.info("Setting chart data...")
        print("Setting chart data...")

        # 차트 데이터 생성
        x_values = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # 첫 번째 열을 X축 값으로 설정
        for i in range(2, ws.max_column + 1):
            y_values = Reference(ws, min_col=i, min_row=2, max_row=ws.max_row)  # 나머지 열들을 Y축 값으로 설정
    
            # 차트에 시리즈 추가
            chart.add_data(y_values, titles_from_data=True)
            chart.set_categories(x_values)

        logging.info("Adding chart to new sheet...")
        print("Adding chart to new sheet...")
        chart_sheet = wb.create_sheet(title="Chart")
        chart_sheet.add_chart(chart, "A1")

        logging.info(f"Saving workbook: {file_path}")
        print(f"Saving workbook: {file_path}")
        wb.save(file_path)

    except Exception as e:
        logging.error(f"Error adding chart to file {file_path}: {e}")
        print(f"Error adding chart to file {file_path}: {e}")
        
def plot_and_save_dynamic_graph(all_data, destination_folder, base_name, title, sort_criteria):
    logging.info("Plotting dynamic graph...")
    print("Plotting dynamic graph...")

    try:
        # 숫자형 열 구분
        numeric_columns = all_data.select_dtypes(include='number').columns.tolist()

        # X축으로 사용한 정렬 기준 열은 Y축에서 제외
        if sort_criteria in numeric_columns:
            numeric_columns.remove(sort_criteria)

        # 만약 Date가 있을 경우 Date를 오른쪽 Y축에 사용할 수 있도록 처리
        if 'Date' in all_data.columns:
            all_data['Date'] = pd.to_datetime(all_data['Date'])  # Date를 datetime으로 변환
            right_y_column = 'Date'
        else:
            right_y_column = None

        # 정렬 기준을 X축으로 사용
        x_axis_column = sort_criteria

        # 만약 X축이 Date인 경우 오른쪽 Y축을 비활성화
        if x_axis_column == 'Date':
            right_y_column = None  # 오른쪽 Y축을 비활성화

        # scatter 그래프 생성
        fig = px.scatter()

        # 숫자형 열들을 왼쪽 Y축에 추가, X축은 sort_criteria 또는 Date
        for column in numeric_columns:
            fig.add_scatter(x=all_data[x_axis_column], y=all_data[column], mode='lines', name=column, yaxis='y1')

        # Date 열을 오른쪽 Y축에 추가 (만약 Date가 있고, X축이 Date가 아닐 경우)
        if right_y_column and x_axis_column != 'Date':
            fig.add_scatter(x=all_data[x_axis_column], y=all_data[right_y_column], mode='lines', name='Date', yaxis='y2')

        # 이중 축 설정 (오른쪽 Y축이 활성화된 경우)
        if right_y_column and x_axis_column != 'Date':
            fig.update_layout(
                title=title,
                xaxis=dict(title=x_axis_column),  # X축: Date 또는 정렬 기준 열
                yaxis=dict(title="Numeric Values", side='left'),  # 왼쪽 Y축: 숫자형 데이터
                yaxis2=dict(title="Date", overlaying='y', side='right')  # 오른쪽 Y축: Date (만약 있을 경우)
            )
        else:
            # 오른쪽 Y축을 비활성화한 레이아웃
            fig.update_layout(
                title=title,
                xaxis=dict(title=x_axis_column),  # X축: Date 또는 정렬 기준 열
                yaxis=dict(title="Numeric Values")  # 왼쪽 Y축만 활성화
            )

        # 결과 파일 저장
        dynamic_chart_file = os.path.join(destination_folder, f'{base_name}_dynamic_chart.html')
        logging.info(f"Saving dynamic chart to: {dynamic_chart_file}")
        print(f"Saving dynamic chart to: {dynamic_chart_file}")
        
        fig.write_html(dynamic_chart_file)
        fig.show()

        return dynamic_chart_file

    except Exception as e:
        logging.error(f"Error plotting dynamic graph: {e}")
        print(f"Error plotting dynamic graph: {e}")
        return None

def merge_excel_files(source_files, destination_folder, sort_criteria, progress_bar, progress_label, root):
    try:
        all_data = pd.DataFrame()
        excluded_sheets = []
        total_steps = len(source_files) + 2  # 파일 처리 + Excel 저장 + HTML 저장

        logging.info("Starting to merge files based on Line...")
        for idx, file_path in enumerate(source_files, start=1):
            logging.info(f"Processing file: {file_path}")
            try:
                xls = pd.ExcelFile(file_path)
                for sheet_name in xls.sheet_names:
                    logging.info(f"Reading sheet: {sheet_name}")
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    if 'Line' in df.columns and pd.api.types.is_numeric_dtype(df['Line']):
                        df = df.set_index('Line')
                        all_data = pd.concat([all_data, df])
                    else:
                        logging.info(f"Excluding sheet: {sheet_name} because it does not contain 'Line' or data is not numeric")
                        excluded_sheets.append(f"{sheet_name} in {os.path.basename(file_path)}")
            except Exception as e:
                logging.error(f"Error processing file {file_path}: {e}")
            
            # Update progress bar
            progress = int((idx / total_steps) * 100)
            progress_bar['value'] = progress
            progress_label.config(text=f"진행률: {progress}%")
            root.update_idletasks()

        # Sort by Line and save the initial merge result
        all_data = all_data.sort_index().groupby(level=0).last().reset_index()  # 중복된 Line 값 처리
        base_name = "merged_file"  # base_name을 간단하게 설정
        timestamp = time.strftime('%Y%m%d_%H%M%S')

        # 간단한 파일 이름 생성
        result_file = generate_simple_filename(base_name, 1)
        result_file = os.path.join(destination_folder, result_file)
        all_data.to_excel(result_file, index=False)
        
        logging.info(f"Saved merged file to: {result_file}")

        # Check if the user-specified sort column exists
        logging.info(f"Reloading merged file for sorting by {sort_criteria}...")
        if sort_criteria in all_data.columns:
            all_data = all_data.sort_values(by=sort_criteria, na_position='last')

            # Fill any missing values in the sort column
            all_data[sort_criteria] = all_data[sort_criteria].fillna(method='ffill')
            all_data[sort_criteria] = all_data[sort_criteria].fillna(method='bfill')

            # Move the sort column to the front (first column)
            cols = all_data.columns.tolist()
            cols.insert(0, cols.pop(cols.index(sort_criteria)))
            all_data = all_data[cols]

            # 간단한 파일 이름 생성
            result_file_sorted = generate_simple_filename(base_name, 2)
            result_file_sorted = os.path.join(destination_folder, result_file_sorted)
            all_data.to_excel(result_file_sorted, index=False)

            logging.info(f"Saved sorted file to: {result_file_sorted}")
            
            # Define chart title here to avoid 'not defined' errors
            chart_title = f"Chart based on {sort_criteria}"

            # Add chart to the sorted Excel file
            logging.info("Adding chart to Excel file...")
            add_chart_to_excel(result_file_sorted)

            # Create dynamic chart based on sorted data
            logging.info("Creating dynamic chart from merged file...")
            dynamic_chart_file_sort = plot_from_merged_excel_twice(result_file, result_file_sorted, destination_folder, base_name, chart_title)
            
        else:
            # Column not found, issue a warning
            logging.warning(f"Column {sort_criteria} not found in the data.")
            root.after(0, lambda: messagebox.showwarning('경고', f'열 {sort_criteria}을(를) 찾을 수 없습니다. 프로그램을 종료합니다.'))
            root.after(0, root.quit)

        progress_bar['value'] = 100
        progress_label.config(text="진행률: 100%")
        root.update_idletasks()

        # Log final results
        logging.info(f"Final merged and sorted file saved: {result_file_sorted}")
        logging.info(f"Dynamic chart saved: {dynamic_chart_file_sort}")

    except Exception as e:
        logging.error(f"Error during merging process: {str(e)}")
        root.after(0, lambda: messagebox.showerror("오류", f"오류가 발생했습니다: {str(e)}"))

def plot_from_merged_excel_twice(line_file_path, sorted_file_path, destination_folder, base_name, title):
    logging.info(f"Loading merged Excel file based on Line: {line_file_path}")
    print(f"Loading merged Excel file based on Line: {line_file_path}")
    
    try:
        line_data = pd.read_excel(line_file_path)
        numeric_columns = line_data.select_dtypes(include='number').columns.tolist()
        x_axis_column = line_data.columns[0]
        
        if 'Line' in numeric_columns:
            numeric_columns.remove('Line')
        if 'Date' in line_data.columns:
            line_data['Date'] = pd.to_datetime(line_data['Date'], errors='coerce')
            right_y_column = 'Date'
        else:
            right_y_column = None

        fig = px.scatter()
        
        for column in numeric_columns:
            fig.add_scatter(x=line_data[x_axis_column], y=line_data[column], mode='lines', name=column, yaxis='y1')
        
        if right_y_column:
            fig.add_scatter(x=line_data[x_axis_column], y=line_data[right_y_column], mode='lines', name='Date', yaxis='y2')
            fig.update_layout(
                title=title,  # 여기에서 title 적용
                xaxis=dict(title=x_axis_column),
                yaxis=dict(title="Numeric Values", side='left'),
                yaxis2=dict(title="Date", overlaying='y', side='right')
            )
        else:
            fig.update_layout(
                title=title,  # 여기에서 title 적용
                xaxis=dict(title=x_axis_column),
                yaxis=dict(title="Numeric Values")
            )
        
        dynamic_chart_file_line = os.path.join(destination_folder, f'{base_name}_line_chart.html')
        fig.write_html(dynamic_chart_file_line)
        fig.show()
        logging.info(f"Dynamic chart file saved: {dynamic_chart_file_line}")

    except Exception as e:
        logging.error(f"Error plotting dynamic graph based on Line: {e}")
        print(f"Error plotting dynamic graph based on Line: {e}")

    # 두 번째 그래프에도 동일하게 적용
    try:
        sorted_data = pd.read_excel(sorted_file_path)
        numeric_columns = sorted_data.select_dtypes(include='number').columns.tolist()
        x_axis_column = sorted_data.columns[0]
        
        if 'Line' in numeric_columns:
            numeric_columns.remove('Line')
        if 'Date' in sorted_data.columns:
            sorted_data['Date'] = pd.to_datetime(sorted_data['Date'], errors='coerce')
            right_y_column = 'Date'
        else:
            right_y_column = None

        fig = px.scatter()
        
        for column in numeric_columns:
            fig.add_scatter(x=sorted_data[x_axis_column], y=sorted_data[column], mode='lines', name=column, yaxis='y1')
        
        if right_y_column:
            fig.add_scatter(x=sorted_data[x_axis_column], y=sorted_data[right_y_column], mode='lines', name='Date', yaxis='y2')
            fig.update_layout(
                title=title,  # 여기에서 title 적용
                xaxis=dict(title=x_axis_column),
                yaxis=dict(title="Numeric Values", side='left'),
                yaxis2=dict(title="Date", overlaying='y', side='right')
            )
        else:
            fig.update_layout(
                title=title,  # 여기에서 title 적용
                xaxis=dict(title=x_axis_column),
                yaxis=dict(title="Numeric Values")
            )

        # Save the sorted chart file
        dynamic_chart_file_sort = os.path.join(destination_folder, f'{base_name}_sorted_chart.html')
        fig.write_html(dynamic_chart_file_sort)
        fig.show()
        logging.info(f"Dynamic chart file (sorted) saved: {dynamic_chart_file_sort}")

    except Exception as e:
        logging.error(f"Error plotting dynamic graph based on Sort: {e}")
        print(f"Error plotting dynamic graph based on Sort: {e}")

def plot_and_save_dual_axis_graph(all_data, destination_folder, base_name, title):
    logging.info("Plotting dual-axis graph with Date on the right Y-axis...")
    print("Plotting dual-axis graph with Date on the right Y-axis...")
    
    try:
        fig = px.scatter()

        # Add numeric variables to the left Y-axis
        numeric_columns = all_data.select_dtypes(include=[np.number]).columns.tolist()
        for column in numeric_columns:
            fig.add_scatter(x=all_data.index, y=all_data[column], mode='lines', name=column, yaxis='y1')

        # Add the Date column to the right Y-axis
        if 'Date' in all_data.columns:
            fig.add_scatter(x=all_data.index, y=all_data['Date'], mode='lines', name='Date', yaxis='y2')

        # Update layout for dual axis
        fig.update_layout(
            title=title,
            xaxis=dict(title="Index"),
            yaxis=dict(title="Numeric Values", side='left'),
            yaxis2=dict(title="Date", overlaying='y', side='right')
        )

        dynamic_chart_file = os.path.join(destination_folder, f'{base_name}_dual_axis_chart.html')
        logging.info(f"Saving dual-axis chart to: {dynamic_chart_file}")
        print(f"Saving dual-axis chart to: {dynamic_chart_file}")
        
        fig.write_html(dynamic_chart_file)
        fig.show()

        return dynamic_chart_file

    except Exception as e:
        logging.error(f"Error plotting dual-axis graph: {e}")
        print(f"Error plotting dual-axis graph: {e}")
        return None



def select_files(title):
    root = Tk()
    root.withdraw()
    messagebox.showinfo('파일 선택', '여러 파일을 선택하려면 Ctrl 또는 Shift 키를 누르고 선택하세요.')
    files_selected = filedialog.askopenfilenames(title=title, filetypes=[("Excel files", "*.xlsx")])
    if not files_selected:
        messagebox.showwarning('경고', '파일을 선택하지 않았습니다.')
        exit()
    logging.info(f"Selected files: {files_selected}")
    print(f"Selected files: {files_selected}")
    return files_selected

def select_folder(title):
    root = Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory(title=title)
    if not folder_selected:
        messagebox.showwarning('경고', '폴더를 선택하지 않았습니다.')
        exit()
    logging.info(f"Selected folder: {folder_selected}")
    print(f"Selected folder: {folder_selected}")
    return folder_selected

def select_sort_criteria():
    root = Tk()
    root.withdraw()
    sort_criteria = simpledialog.askstring("정렬 기준", "정렬 기준을 입력하세요 (정렬 기준열 이름을 그대로 입력):")
    if not sort_criteria:
        messagebox.showwarning('경고', '정렬 기준을 입력하지 않았습니다.')
        exit()
    logging.info(f"Selected sort criteria: {sort_criteria}")
    print(f"Selected sort criteria: {sort_criteria}")
    return sort_criteria

def on_closing(root):
    root.destroy()
    sys.exit()

def run_merge_process(source_files, destination_folder, sort_criteria, progress_bar, progress_label, root):
    threading.Thread(target=merge_excel_files, args=(source_files, destination_folder, sort_criteria, progress_bar, progress_label, root)).start()

def setup_logging():
    print("Setting up logging...")  # 디버깅을 위해 print 추가
    logging.basicConfig(filename='merge_excel_log.txt', level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s')
    logging.info('Logging setup complete.')


# 로그 초기화를 메인 함수 이전에 수행하여 모든 오류를 캡처
setup_logging()

def main():
    try:
        setup_logging()
        
        logging.info('Program started.')

        logging.info('병합할 엑셀 파일들을 선택하세요 / Select Excel files to merge')
        print('병합할 엑셀 파일들을 선택하세요 / Select Excel files to merge')

        # Use a single root for all Tkinter dialogs
        root = Tk()
        root.withdraw()  # Hide the root window initially

        source_files = select_files('병합할 엑셀 파일들을 선택하세요 / Select Excel files to merge')
        if not source_files:
            logging.error('No files selected.')
            print("No files selected.")
            return
        
        logging.info('대상 폴더를 선택하세요 / Select destination folder')
        print('대상 폴더를 선택하세요 / Select destination folder')
        destination_folder = select_folder('대상 폴더를 선택하세요 / Select destination folder')
        if not destination_folder:
            logging.error('No destination folder selected.')
            print("No destination folder selected.")
            return

        sort_criteria = select_sort_criteria()
        if not sort_criteria:
            logging.error('No sort criteria provided.')
            print("No sort criteria provided.")
            return

        # 중복 데이터 처리 알림
        messagebox.showinfo('중복 데이터 처리', '중복된 데이터는 마지막 값으로 처리됩니다.')

        # Show the main application window for merging process
        root.deiconify()
        root.title("Merging Progress")

        progress_label = ttk.Label(root, text="진행률: 0%")
        progress_label.grid(row=0, column=0, padx=10, pady=10)

        progress_bar = ttk.Progressbar(root, length=300, mode='determinate')
        progress_bar.grid(row=1, column=0, padx=10, pady=10)

        # 창 닫기 이벤트 처리
        root.protocol("WM_DELETE_WINDOW", lambda: on_closing(root))

        # 병합 프로세스를 별도의 스레드에서 실행
        run_merge_process(source_files, destination_folder, sort_criteria, progress_bar, progress_label, root)

        root.mainloop()
    
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        print(f"An error occurred: {str(e)}")
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

    finally:
        # 프로그램이 종료되지 않고 콘솔 창에서 대기하도록 설정
        print("프로그램이 종료되지 않고 대기 중입니다. 콘솔 창을 닫으려면 아무 키나 누르세요...")
        input("Press Enter to exit...")

if __name__ == "__main__":
    main()
